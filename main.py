from Function.classScanQR import DiemDanh
from Function.classDatabase import *
from pages.add_student import AddStudent
from pages.setting_cam import SettingCamera
from ui.main_window import Ui_MainWindow
from Function.loadStudent import LoadStudent
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from tkinter import messagebox
from QRCode.create import *
import pandas
import os, json, subprocess
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import sys, re
from PyQt5.QtWidgets import QApplication, QFileDialog

def select_excel_file():
    dialog = QFileDialog()
    dialog.setNameFilter("Excel files (*.xlsx)")
    dialog.setWindowTitle("Select Excel File")
    dialog.setFileMode(QFileDialog.ExistingFile)

    if dialog.exec_() == QFileDialog.Accepted:
        file_path = dialog.selectedFiles()[0]
        return os.path.abspath(file_path)


my_setting = QSettings("Điểm danh THCS", 'Setting1')

class MainApp(QMainWindow, Ui_MainWindow):
    thread_count_load = 0
    cam_index = 0
    show_cam = False
    data = {
        'show_cam': 'Show camera screen'
    }
    # student_json = my_setting.value('student_json', {})
    threadScan = None
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.icon = QIcon(os.path.join(os.getcwd(), 'assert', 'icon', 'logo.png'))
        self.setWindowIcon(self.icon)
        self.setting_cam_btn.clicked.connect(self.show_setting)
        self.diem_danh_btn.clicked.connect(self.started)
        self.add_student_btn.clicked.connect(self.add_student)
        self.view_result_btn.clicked.connect(self.view_result)
        self.export_qrCode_btn.clicked.connect(self.export_qr)
        # self.load_student()
        self.view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.view.customContextMenuRequested.connect(self.menu)
        self.load_student()
        self.excel_btn.clicked.connect(self.import_excel)
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.load_student)
        self.timer.start(15000)
    
    def import_excel(self):
        filepath = select_excel_file()
        if filepath:
            dataSet = []
            df = pandas.read_excel(filepath)
            for index, row in df.iterrows():
                data = json.loads(f"{row.to_json()}")
                data_ = []
                for k, v in dict(data).items():
                    data_.append(v)
                dataSet.append(data_)
            success, error = 0, 0, 
            for data in dataSet:
                name, msv = data[0], data[1]
                if db.query(DB_STUDENT, '*', f'`msv` = "{msv}"') == []:
                    db.insert(DB_STUDENT, '(`name`, `msv`, `dateEnd`)', f'("{name}", "{msv}", {round(time.time())})')
                    row = self.view.rowCount()
                    self.add_row(self.view, row)
                    self.view.setRowHeight(row, 8)
                    for x, y in zip(range(6), [row+1, name, msv, '', '', '']):
                        self.edit_row(self.view, row, x, y)
                    qr_code = create_qr(f'{name}|{msv}')
                    if qr_code: success+=1
                    else: error += 1
            messagebox.showinfo("Thông báo", "Xong !\nSuccess: {}\nError: {}".format(success, error))
            self.load_student()
    def isHourFormat(self, string: str):
        if re.match("\d{2}:\d{2}:\d{2}", string):
            try:
                datetime.strptime(string, "%H:%M:%S")
                return True
            except: pass
            
    def view_result(self):
        name_data = []
        msv_data = []
        time_in_data = []
        time_out_data = []
        status_data = []
        item = self.view.item
        for row in range(self.view.rowCount()):
            time_unix_in = item(row, 3).text()
            time_unix_out = item(row, 4).text()
            
            time_in = time_unix_in if self.isHourFormat(time_unix_in) else "Chưa có"
            time_out = time_unix_out if self.isHourFormat(time_unix_out) else "Chưa có"
            name_data.append(item(row, 1).text())
            msv_data.append(item(row, 2).text())
            time_in_data.append(time_in)
            time_out_data.append(time_out)
            status_data.append(item(row, 5).text())

        # Set data thành json để format sang xlsx bằng pandas
        data = {
            "Họ và tên": name_data,
            "Mã sinh viên": msv_data,
            "Thời gian vào": time_in_data,
            "Thời gian ra": time_out_data,
            "Kết quả": status_data,
        }
        
        # Sử dụng os.path.join để tạo đường dẫn file
        file_name = '{}.xlsx'.format(datetime.fromtimestamp(int(time.time())).strftime("%d-%m-%Y"))
        folder_path = os.path.join(os.getcwd(), 'output')
        file_path = os.path.join(folder_path, file_name)
        
        # Tạo thư mục nếu nó không tồn tại
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        
        # Tạo và lưu DataFrame vào Excel
        data_excel = pandas.DataFrame(data)
        data_excel.to_excel(file_path, index=False)
        
        # Tải workbook với openpyxl và điều chỉnh độ rộng cột
        workbook = load_workbook(file_path)
        sheet = workbook.active
        for column_cells in sheet.columns:
            length = max(len(self.as_text(cell.value)) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 30

        center_aligned_text = Alignment(horizontal='center', vertical='center')

        # Áp dụng đối tượng Alignment cho mỗi cell trong sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = center_aligned_text
                
        # Lưu lại các thay đổi vào workbook
        workbook.save(file_path)
        
        # Mở thư mục và chọn file trên Windows
        if os.name == 'nt':
            subprocess.Popen(f'explorer /select,"{file_path}"')
            
    def as_text(self, value):
        if value is None:
            return ""
        return str(value)

    def export_qr(self):
        # Tạo workbook mới và lấy sheet đầu tiên
        wb = Workbook()
        ws = wb.active

        # Đặt tiêu đề cho các cột
        ws['A1'] = 'Họ và Tên'
        ws['B1'] = 'Mã QR'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        data = []
        item = self.view.item
        for row in range(self.view.rowCount()):
            data.append((item(row, 1).text(), f'{item(row, 2).text()}.png'))
        for row, (name, qr_path) in enumerate(data, start=2):
            qr_path = '{}\\{}'.format(os.path.join(os.getcwd(), 'imgQrCode'), qr_path)
            
            cell = ws.cell(row=row, column=1)
            cell.value = name
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Chèn hình ảnh QR code và căn chỉnh ở giữa
            img = Image(qr_path)
            img.width, img.height = 120, 120
            img.left = 15
            img.anchor = f'B{row}'
            # Căn chỉnh hình ảnh ở giữa cell
            ws.add_image(img)

            ws.row_dimensions[row].height = 120 
            ws.column_dimensions['B'].width = 60 
            ws.column_dimensions['A'].width = 30
        
        file_name = 'data_qrCode.xlsx'
        file_path = os.path.join(os.getcwd(), file_name)
        wb.save(file_path)
        if os.name == 'nt':
            subprocess.Popen(f'explorer /select,"{file_path}"')
        
    def menu(self):
        menu = QMenu()
        delete = menu.addAction("Xóa học sinh")
        show_cam = menu.addAction(self.data['show_cam'])
        action = menu.exec_(QCursor.pos())
        if action == show_cam:
            if show_cam.text() == "Show camera screen":
                self.show_cam = True
                self.data['show_cam'] = "Hide camera screen"
            else:
                self.show_cam = False
                self.data['show_cam'] = "Show camera screen"
        if action == delete:
            table = self.sender()
            selected = table.selectionModel().selectedRows()
            for row in selected:
                try:
                    row = row.row()
                    msv = table.item(row, 2).text()
                    db.delete(DB_STUDENT, '`msv` = "{}"'.format(msv))
                    os.unlink('./imgQrCode/{}.png'.format(msv))
                except:pass
            self.load_student()
        
    def show_setting(self):
        setting_cam = SettingCamera()
        setting_cam.setWindowIcon(self.icon)
        result = setting_cam.exec_()
        if result == 1:
            cam_index = setting_cam.index_cam
            if cam_index: self.cam_index = cam_index
    
    def load_student(self):
        self.thread_count_load = 0
        try:
            self.view.itemChanged.disconnect()
        except: pass
        data = db.query(DB_STUDENT, '*')
        self.view.setRowCount(0)
        if data != []:
            for _ in data:self.thread_count_load+=1
            for x, y in enumerate(data):
                    row = self.view.rowCount()
                    self.view.insertRow(row)
                    self.oo = LoadStudent(self, row, y)
                    self.oo.edit_row.connect(self.edit_row)
                    self.oo.finished.connect(self.load_finished)
                    self.oo.start()
                    self.view.setRowHeight(row, 8)
                    self.sleep(0.002)
    
    def load_finished(self):
        self.thread_count_load -= 1
        if self.thread_count_load == 0:
            self.view.itemChanged.connect(self.dataChange)
    
    def dataChange(self, item: QTableWidgetItem):
        if item.column() not in [0, 2]:
            try:
                text = item.text()
                col  = item.column()
                row = item.row()
                msv = self.view.item(row, 2).text()
                if col == 1: key = 'name'
                elif col == 3: key = 'time_in'
                elif col == 4: key = 'time_out'
                elif col == 5: key = 'status'
                db.update(DB_STUDENT, f'`{key}` = "{text}"', f'`msv` = "{msv}"')
            except Exception as e: print(e)
            
    def add_student(self):
        try:self.view.itemChanged.disconnect()
        except: pass
        add_st = AddStudent()
        add_st.setWindowIcon(self.icon)
        if add_st.exec_() == 1:
            name, msv = add_st.name, add_st.msv
            if db.query(DB_STUDENT, '*', f'`msv` = "{msv}"') == []:
                db.insert(DB_STUDENT, '(`name`, `msv`, `dateEnd`)', f'("{name}", "{msv}", {round(time.time())})')
                row = self.view.rowCount()
                self.add_row(self.view, row)
                self.view.setRowHeight(row, 8)
                for x, y in zip(range(6), [row+1, name, msv, '', '', '']):
                    self.edit_row(self.view, row, x, y)
                qr_code = create_qr(f'{name}|{msv}')
                if qr_code: messagebox.showinfo("Thông báo", "Tạo QR Code thành công\nSave tại: {}".format(qr_code))
            else:
                if messagebox.askyesno("Xảy ra lỗi", "Mã sinh viên này đã tồn tại\nCập nhật thông tin mới ?"):
                    db.update(DB_STUDENT, f'`name` = "{name}"', f'`msv` = "{msv}"')
        self.view.itemChanged.connect(self.dataChange)
    
    def edit_row(self, table: QTableWidget, row: int, column: int, text: str):
        item = QTableWidgetItem(str(text))
        if column == 0: item.setTextAlignment(Qt.AlignCenter)
        item.setToolTip(str(text))
        table.setItem(row, column, item)
        
    def add_row(self, table: QTableWidget, r: int):
        table.insertRow(r)    
            
    def started(self):
        if self.diem_danh_btn.text() == "Bắt đầu điểm danh":
            self.diem_danh_btn.setText("Dừng điểm danh")
            self.setting_cam_btn.setEnabled(False)
            self.add_student_btn.setEnabled(False)
            self.view_result_btn.setEnabled(False)
            self.export_qrCode_btn.setEnabled(False)
            self.threadScan = DiemDanh(self)
            self.threadScan.loadAccount.connect(self.load_student)
            self.threadScan.start()
            self.sleep(0.05)
        else:
            self.diem_danh_btn.setText("Bắt đầu điểm danh")
            self.setting_cam_btn.setEnabled(True)
            self.add_student_btn.setEnabled(True)
            self.view_result_btn.setEnabled(True)
            self.export_qrCode_btn.setEnabled(True)
            try:
                self.threadScan.stop()
            except: pass
            
    def sleep(self, time: int):
        delay = QEventLoop(self)
        QTimer.singleShot(int(time * 1000), delay.quit)
        delay.exec_()
if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)
    dialog = MainApp()
    dialog.show()
    sys.exit(app.exec_())